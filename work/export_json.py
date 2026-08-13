import json, os, csv
from datetime import datetime
import openpyxl

XLSX_SOURCE = r"C:\Users\ASUS\Downloads\lets-make-a-dashboard-using-this (2)\lets-make-a-dashboard-using-this\Untitled spreadsheet121.xlsx"
CSV_SOURCE = r"C:\Users\ASUS\Downloads\lets-make-a-dashboard-using-this (2)\lets-make-a-dashboard-using-this\company_master_data_2026-08-12.csv"
OUTPUT = r"C:\Users\ASUS\Downloads\lets-make-a-dashboard-using-this (2)\lets-make-a-dashboard-using-this\lets-make-a-dashboard-using-this\outputs\company-dashboard\companies-full.json"

def clean(value):
    return str(value or "").strip()

def parse_founders(people_raw, links_raw):
    founders = []
    people = clean(people_raw).split("\n")
    links = clean(links_raw).split(",\n")
    for i, person in enumerate(people):
        if not person.strip(): continue
        parts = person.split(";")
        name = clean(parts[0]) if len(parts) > 0 else ""
        title = clean(parts[1]) if len(parts) > 1 else ""
        linkedin = clean(links[i]) if i < len(links) else ""
        if name:
            founders.append({"name": name, "title": title, "linkedin": linkedin})
    return founders

def parse_city_from_address(address):
    parts = [p.strip() for p in address.split(',')]
    if len(parts) >= 4:
        return parts[-4]
    elif len(parts) >= 3:
        return parts[-3]
    return ""

def main():
    companies = []
    seen = set()

    wb = openpyxl.load_workbook(XLSX_SOURCE, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))
        name = clean(r.get("Company Name"))
        if not name: continue
        sector = clean(r.get("Sector (Pratice Area & Feed)"))
        desc = clean(r.get("Description")) or clean(r.get("Overview"))
        companies.append({
            "name": name,
            "domain": clean(r.get("Domain Name")),
            "state": clean(r.get("State")),
            "district": clean(r.get("City")),
            "city": clean(r.get("City")),
            "founded": clean(r.get("Founded Year")),
            "sector": sector[:130] + ("\u2026" if len(sector) > 130 else ""),
            "stage": clean(r.get("Company Stage")),
            "funded": clean(r.get("Is Funded")),
            "funding": "",
            "employees": clean(r.get("Total Employee Count")),
            "website": clean(r.get("Website")),
            "linkedin": clean(r.get("LinkedIn")),
            "description": desc[:220] + ("\u2026" if len(desc) > 220 else ""),
            "founders": parse_founders(r.get("Key People Info"), r.get("Links to Key People Profiles"))
        })
        seen.add(name.lower())
    wb.close()
    print(f"Loaded {len(companies)} from xlsx")

    csv_count = 0
    with open(CSV_SOURCE, newline="", encoding="utf-8-sig", errors="replace") as f:
        for row in csv.DictReader(f):
            name = clean(row.get("Company Name"))
            if not name or name.lower() in seen: continue
            state = clean(row.get("Company State"))
            address = clean(row.get("Company Address"))
            city = parse_city_from_address(address)
            sector = clean(row.get("Company Industrial Classification"))
            reg_date = clean(row.get("Company Registration Date"))
            founded = reg_date[:4] if reg_date else ""
            companies.append({
                "name": name, "domain": "", "state": state,
                "district": city, "city": city, "founded": founded,
                "sector": sector, "stage": clean(row.get("Company Class")),
                "funded": "", "funding": "", "employees": "",
                "website": "", "linkedin": "", "description": "", "founders": []
            })
            seen.add(name.lower())
            csv_count += 1
    print(f"Loaded {csv_count} from CSV")

    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump({"generatedAt": datetime.now().isoformat(timespec="seconds"), "companies": companies}, f, ensure_ascii=False, separators=(",", ":"))
    print(f"Total {len(companies)} companies saved to {OUTPUT}")

if __name__ == "__main__": main()
