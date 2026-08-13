import json, os, csv
import openpyxl
from datetime import datetime

XLSX = r"C:\Users\ASUS\Downloads\lets-make-a-dashboard-using-this (2)\lets-make-a-dashboard-using-this\Untitled spreadsheet121.xlsx"
CSV_SOURCE = r"C:\Users\ASUS\Downloads\lets-make-a-dashboard-using-this (2)\lets-make-a-dashboard-using-this\company_master_data_2026-08-12.csv"
OUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "outputs", "company-dashboard")
CHUNK_SIZE = 20000  # companies per chunk

def clean(v): return str(v or "").strip()

def pf(p, l):
    fs = []
    people = clean(p).split("\n")
    links = clean(l).split(",\n")
    for i, x in enumerate(people):
        if not x.strip(): continue
        pts = x.split(";")
        n = clean(pts[0]); t = clean(pts[1]) if len(pts) > 1 else ""
        lk = clean(links[i]) if i < len(links) else ""
        if n: fs.append({"name": n, "title": t, "linkedin": lk})
    return fs

def parse_city(address):
    parts = [p.strip() for p in address.split(',')]
    return parts[-4] if len(parts) >= 4 else (parts[-3] if len(parts) >= 3 else "")

# Load all companies
companies = []
seen = set()

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb.active
hdrs = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
for row in ws.iter_rows(min_row=2, values_only=True):
    r = dict(zip(hdrs, row))
    n = clean(r.get("Company Name"))
    if not n: continue
    s = clean(r.get("Sector (Pratice Area & Feed)"))
    d = clean(r.get("Description")) or clean(r.get("Overview"))
    companies.append({"name": n, "domain": clean(r.get("Domain Name")), "state": clean(r.get("State")), "district": clean(r.get("City")), "city": clean(r.get("City")), "founded": clean(r.get("Founded Year")), "sector": s[:130] + ("\u2026" if len(s) > 130 else ""), "stage": clean(r.get("Company Stage")), "funded": clean(r.get("Is Funded")), "funding": "", "employees": clean(r.get("Total Employee Count")), "website": clean(r.get("Website")), "linkedin": clean(r.get("LinkedIn")), "description": d[:220] + ("\u2026" if len(d) > 220 else ""), "founders": pf(r.get("Key People Info"), r.get("Links to Key People Profiles"))})
    seen.add(n.lower())
wb.close()
print(f"xlsx: {len(companies)}")

with open(CSV_SOURCE, newline="", encoding="utf-8-sig", errors="replace") as f:
    for row in csv.DictReader(f):
        n = clean(row.get("Company Name"))
        if not n or n.lower() in seen: continue
        state = clean(row.get("Company State"))
        city = parse_city(clean(row.get("Company Address")))
        sector = clean(row.get("Company Industrial Classification"))
        founded = clean(row.get("Company Registration Date"))[:4]
        companies.append({"name": n, "domain": "", "state": state, "district": city, "city": city, "founded": founded, "sector": sector, "stage": clean(row.get("Company Class")), "funded": "", "funding": "", "employees": "", "website": "", "linkedin": "", "description": "", "founders": []})
        seen.add(n.lower())
print(f"total: {len(companies)}")

# Write chunk files
chunks = [companies[i:i+CHUNK_SIZE] for i in range(0, len(companies), CHUNK_SIZE)]
for idx, chunk in enumerate(chunks):
    path = os.path.join(OUT_DIR, f"data-chunk-{idx}.js")
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"window.__CHUNK_{idx} = " + json.dumps(chunk, ensure_ascii=False, separators=(",", ":")) + ";")
    size_mb = os.path.getsize(path) / 1024 / 1024
    print(f"chunk {idx}: {len(chunk)} companies, {size_mb:.1f}MB -> {path}")

# Write manifest
manifest_path = os.path.join(OUT_DIR, "data-manifest.js")
with open(manifest_path, "w") as f:
    f.write(f"window.__DATA_CHUNKS = {len(chunks)};")
print(f"manifest written: {len(chunks)} chunks")
