import json, os
from datetime import datetime
import openpyxl

XLSX = r"C:\Users\ASUS\Downloads\lets-make-a-dashboard-using-this (2)\lets-make-a-dashboard-using-this\Untitled spreadsheet121.xlsx"
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "outputs", "company-dashboard", "dashboard-data.js")

def clean(v): return str(v or "").strip()

def pf(p, l):
    fs = []
    people = clean(p).split("\n")
    links = clean(l).split(",\n")
    for i, x in enumerate(people):
        if not x.strip(): continue
        pts = x.split(";")
        n = clean(pts[0]); t = clean(pts[1]) if len(pts) > 1 else ""
        lk = clean(links[i]) if i < len(links) else ""
        if n: fs.append({"name": n, "title": t, "linkedin": lk})
    return fs

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb.active
hdrs = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
cos = []
for row in ws.iter_rows(min_row=2, values_only=True):
    r = dict(zip(hdrs, row))
    n = clean(r.get("Company Name"))
    if not n: continue
    s = clean(r.get("Sector (Pratice Area & Feed)"))
    d = clean(r.get("Description")) or clean(r.get("Overview"))
    cos.append({"name": n, "domain": clean(r.get("Domain Name")), "state": clean(r.get("State")), "district": clean(r.get("City")), "city": clean(r.get("City")), "founded": clean(r.get("Founded Year")), "sector": s[:130] + ("\u2026" if len(s) > 130 else ""), "stage": clean(r.get("Company Stage")), "funded": clean(r.get("Is Funded")), "funding": "", "employees": clean(r.get("Total Employee Count")), "website": clean(r.get("Website")), "linkedin": clean(r.get("LinkedIn")), "description": d[:220] + ("\u2026" if len(d) > 220 else ""), "founders": pf(r.get("Key People Info"), r.get("Links to Key People Profiles"))})
wb.close()
open(OUT, "w", encoding="utf-8").write("window.DASHBOARD_DATA = " + json.dumps({"generatedAt": datetime.now().isoformat(timespec="seconds"), "companies": cos}, ensure_ascii=False, separators=(",", ":")) + ";")
print(f"Wrote {len(cos)} companies")
